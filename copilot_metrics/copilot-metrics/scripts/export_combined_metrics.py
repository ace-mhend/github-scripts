#!/usr/bin/env python3
"""
============================================
Export Combined Copilot Metrics to Excel
============================================
Combines user data with aggregate metrics
Sorted by username
Includes chat requests and lines of code
============================================
"""

import argparse
import glob
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def find_latest_file(pattern):
    """Find the most recently modified file matching a glob pattern."""
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def parse_datetime(dt_str):
    """Parse ISO datetime string to datetime object."""
    if not dt_str:
        return None
    try:
        # Handle various ISO formats
        dt_str = dt_str.replace('Z', '+00:00')
        if '.' in dt_str:
            return datetime.fromisoformat(dt_str.split('.')[0])
        return datetime.fromisoformat(dt_str.replace('Z', ''))
    except (ValueError, TypeError):
        return None


def add_table_style(ws, data_range, table_name):
    """Add a table style to a range of cells."""
    table = Table(displayName=table_name, ref=data_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def main():
    parser = argparse.ArgumentParser(description='Export Copilot metrics to Excel')
    parser.add_argument('--output-dir', default='./output', help='Output directory')
    parser.add_argument('--days-back', type=int, default=1, help='Days back for activity filter')
    args = parser.parse_args()

    output_dir = os.path.abspath(args.output_dir)
    
    # Find the most recent combined files
    users_file = find_latest_file(os.path.join(output_dir, 'combined_users_*.json'))
    metrics_file = find_latest_file(os.path.join(output_dir, 'combined_metrics_*.json'))

    if not users_file or not os.path.exists(users_file):
        print("ERROR: No users JSON file found. Run fetch_combined_metrics.sh first.")
        sys.exit(1)

    if not metrics_file or not os.path.exists(metrics_file):
        print("ERROR: No metrics JSON file found. Run fetch_combined_metrics.sh first.")
        sys.exit(1)

    print(f"Processing users: {users_file}")
    print(f"Processing metrics: {metrics_file}")

    # Load JSON data
    with open(users_file, 'r') as f:
        users_json = json.load(f)
    with open(metrics_file, 'r') as f:
        metrics_json = json.load(f)

    cutoff_date = datetime.now() - timedelta(days=args.days_back)

    # ============================================
    # Sheet 1: Users sorted by username with activity
    # ============================================
    user_data = []
    for seat in users_json.get('seats', []):
        assignee = seat.get('assignee', {})
        last_activity = parse_datetime(seat.get('last_activity_at'))
        last_auth = parse_datetime(seat.get('last_authenticated_at'))
        created_at = parse_datetime(seat.get('created_at'))

        active_in_period = "Yes" if last_activity and last_activity > cutoff_date else "No"

        user_data.append({
            'Username': assignee.get('login', 'Unknown'),
            'PlanType': seat.get('plan_type', ''),
            'LastActivityAt': last_activity.isoformat() if last_activity else '',
            'LastActivityEditor': seat.get('last_activity_editor', ''),
            'LastAuthenticatedAt': last_auth.isoformat() if last_auth else '',
            'CreatedAt': created_at.isoformat() if created_at else '',
            'ActiveInPeriod': active_in_period
        })

    user_data.sort(key=lambda x: x['Username'].lower())
    active_users = [u for u in user_data if u['ActiveInPeriod'] == 'Yes']

    print(f"Total users: {len(user_data)}")
    print(f"Users active in past {args.days_back} day(s): {len(active_users)}")

    # ============================================
    # Sheet 2: Chat requests by day
    # ============================================
    chat_data = []
    for day in metrics_json:
        total_chats = 0
        ide_chat_users = 0
        dotcom_chat_users = 0

        # IDE Chat
        ide_chat = day.get('copilot_ide_chat', {})
        if ide_chat.get('editors'):
            for editor in ide_chat['editors']:
                for model in editor.get('models', []):
                    total_chats += model.get('total_chats', 0)
            ide_chat_users = ide_chat.get('total_engaged_users', 0)

        # Dotcom Chat
        dotcom_chat = day.get('copilot_dotcom_chat', {})
        if dotcom_chat.get('models'):
            for model in dotcom_chat['models']:
                total_chats += model.get('total_chats', 0)
            dotcom_chat_users = dotcom_chat.get('total_engaged_users', 0)

        total_active = day.get('total_active_users', 0)
        chats_per_user = round(total_chats / total_active, 1) if total_active > 0 else 0

        chat_data.append({
            'Date': day.get('date', ''),
            'TotalChatRequests': total_chats,
            'IDEChatUsers': ide_chat_users,
            'DotcomChatUsers': dotcom_chat_users,
            'TotalActiveUsers': total_active,
            'ChatsPerActiveUser': chats_per_user
        })

    chat_data.sort(key=lambda x: x['Date'], reverse=True)

    # ============================================
    # Sheet 3: Lines of code by day and language
    # ============================================
    lines_data = []
    for day in metrics_json:
        completions = day.get('copilot_ide_code_completions', {})
        if completions.get('editors'):
            for editor in completions['editors']:
                for model in editor.get('models', []):
                    for lang in model.get('languages', []):
                        total_suggestions = lang.get('total_code_suggestions', 0)
                        total_acceptances = lang.get('total_code_acceptances', 0)
                        acceptance_rate = round((total_acceptances / total_suggestions) * 100, 1) if total_suggestions > 0 else 0

                        lines_data.append({
                            'Date': day.get('date', ''),
                            'Editor': editor.get('name', ''),
                            'Language': lang.get('name', ''),
                            'loc_added_sum': lang.get('total_code_lines_accepted', 0),
                            'loc_deleted_sum': None,  # Not available in API
                            'loc_suggested_to_add_sum': lang.get('total_code_lines_suggested', 0),
                            'agent_edit': None,  # Not available in API
                            'Acceptances': total_acceptances,
                            'Suggestions': total_suggestions,
                            'AcceptanceRate': acceptance_rate
                        })

    lines_data.sort(key=lambda x: (x['Date'], x['Language']), reverse=True)

    # ============================================
    # Sheet 4: Summary - totals for period
    # ============================================
    dates = [d.get('date', '') for d in metrics_json if d.get('date')]
    period_start = min(dates) if dates else ''
    period_end = max(dates) if dates else ''

    total_chats = sum(d['TotalChatRequests'] for d in chat_data)
    total_loc_added = sum(d['loc_added_sum'] or 0 for d in lines_data)
    total_loc_suggested = sum(d['loc_suggested_to_add_sum'] or 0 for d in lines_data)
    avg_chats_per_user = round(sum(d['ChatsPerActiveUser'] for d in chat_data) / len(chat_data), 1) if chat_data else 0
    overall_acceptance = f"{round((total_loc_added / total_loc_suggested) * 100, 1)}%" if total_loc_suggested > 0 else "0%"

    summary = [
        {'Metric': 'Period Start', 'Value': period_start},
        {'Metric': 'Period End', 'Value': period_end},
        {'Metric': 'Total Seats', 'Value': users_json.get('total_seats', len(user_data))},
        {'Metric': f'Active Users (past {args.days_back} day)', 'Value': len(active_users)},
        {'Metric': 'Total Chat Requests', 'Value': total_chats},
        {'Metric': 'Avg Chats Per Active User', 'Value': avg_chats_per_user},
        {'Metric': 'loc_added_sum (Lines Accepted)', 'Value': total_loc_added},
        {'Metric': 'loc_deleted_sum (Lines Removed)', 'Value': 'N/A - Not in API'},
        {'Metric': 'loc_suggested_to_add_sum (Ghost Text)', 'Value': total_loc_suggested},
        {'Metric': 'agent_edit (Agent Mode Lines)', 'Value': 'N/A - Not in API'},
        {'Metric': 'Overall Acceptance Rate', 'Value': overall_acceptance},
    ]

    # ============================================
    # Export to Excel
    # ============================================
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create monthly folder
    month_folder = os.path.join(output_dir, datetime.now().strftime("%Y-%m"))
    os.makedirs(month_folder, exist_ok=True)
    
    excel_path = os.path.join(month_folder, f"copilot_metrics_{datetime.now().strftime('%Y-%m-%d')}_{timestamp}.xlsx")

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    ws_summary.append(['Metric', 'Value'])
    for row in summary:
        ws_summary.append([row['Metric'], row['Value']])
    
    # Style header
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Active Users sheet
    ws_users = wb.create_sheet('ActiveUsers')
    if active_users:
        headers = list(active_users[0].keys())
        ws_users.append(headers)
        for user in active_users:
            ws_users.append([user[h] for h in headers])
        for cell in ws_users[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

    # Chat Requests sheet
    ws_chat = wb.create_sheet('ChatRequests')
    if chat_data:
        headers = list(chat_data[0].keys())
        ws_chat.append(headers)
        for row in chat_data:
            ws_chat.append([row[h] for h in headers])
        for cell in ws_chat[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

    # Lines of Code sheet
    ws_lines = wb.create_sheet('LinesOfCode')
    if lines_data:
        headers = list(lines_data[0].keys())
        ws_lines.append(headers)
        for row in lines_data:
            ws_lines.append([row[h] for h in headers])
        for cell in ws_lines[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

    # Auto-width columns
    for ws in [ws_summary, ws_users, ws_chat, ws_lines]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)

    print("")
    print(f"Excel file created: {excel_path}")
    print("")
    print("Sheets included:")
    print("  1. Summary - Key metrics overview")
    print("  2. ActiveUsers - Users sorted by username")
    print("  3. ChatRequests - Chat requests by active user per day")
    print("  4. LinesOfCode - Lines of code accepted/suggested by language")


if __name__ == '__main__':
    main()
