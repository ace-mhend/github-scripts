#!/usr/bin/env python3
"""
Export Excel files to CSV format.
Converts monthly summary Excel files to CSV for easier data processing.
"""

import argparse
import os
from pathlib import Path
from openpyxl import load_workbook
import csv


def excel_to_csv(excel_path: Path, output_dir: Path = None):
    """Convert an Excel file to CSV (one CSV per sheet)."""
    if output_dir is None:
        output_dir = excel_path.parent
    
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    base_name = excel_path.stem
    
    created_files = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Clean sheet name for filename
        safe_sheet_name = sheet_name.replace(' ', '_').replace('/', '-')
        csv_filename = f"{base_name}_{safe_sheet_name}.csv"
        csv_path = output_dir / csv_filename
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    writer.writerow(row)
        
        created_files.append(csv_path)
        print(f"Created: {csv_path}")
    
    wb.close()
    return created_files


def main():
    parser = argparse.ArgumentParser(description='Export Excel files to CSV')
    parser.add_argument('--output-dir', '-o', required=True,
                        help='Base output directory containing monthly folders')
    parser.add_argument('--month', '-m', required=True,
                        help='Month folder to process (YYYY-MM format)')
    parser.add_argument('--file', '-f', 
                        help='Specific Excel file to convert (optional)')
    args = parser.parse_args()
    
    output_dir = Path(args.output_dir)
    month_dir = output_dir / args.month
    
    if not month_dir.exists():
        print(f"Error: Month directory not found: {month_dir}")
        return 1
    
    if args.file:
        # Convert specific file
        excel_path = Path(args.file)
        if not excel_path.exists():
            print(f"Error: File not found: {excel_path}")
            return 1
        excel_to_csv(excel_path, month_dir)
    else:
        # Convert all Excel files in the month directory
        excel_files = list(month_dir.glob('*.xlsx'))
        if not excel_files:
            print(f"No Excel files found in {month_dir}")
            return 0
        
        print(f"Found {len(excel_files)} Excel file(s) to convert")
        for excel_path in excel_files:
            print(f"\nProcessing: {excel_path.name}")
            excel_to_csv(excel_path, month_dir)
    
    print("\nCSV export complete!")
    return 0


if __name__ == '__main__':
    exit(main())
