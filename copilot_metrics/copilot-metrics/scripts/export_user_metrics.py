#!/usr/bin/env python3
"""
============================================
Export User Metrics to Excel (Past Day)
============================================
Filters for users active in the past day
Sorted by username
============================================
"""

import argparse
import glob
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def find_latest_file(pattern):
    """Find the most recently modified file matching a glob pattern."""
    files = [f for f in glob.glob(pattern) if not f.endswith('.page1') and not f.endswith('.page2')]
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def parse_datetime(dt_str):
    """Parse ISO datetime string to datetime object."""
    if not dt_str:
        return None
    try:
        dt_str = dt_str.replace('Z', '+00:00')
        if '.' in dt_str:
            return datetime.fromisoformat(dt_str.split('.')[0])
        return datetime.fromisoformat(dt_str.replace('Z', ''))
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser(description='Export user metrics to Excel')
    parser.add_argument('--input-file', help='Input JSON file (optional, uses latest)')
    parser.add_argument('--output-dir', default='./output', help='Output directory')
    args = parser.parse_args()

    output_dir = os.path.abspath(args.output_dir)
    
    # Find the most recent user metrics file if not specified
    input_file = args.input_file
    if not input_file:
        input_file = find_latest_file(os.path.join(output_dir, 'user_metrics_*.json'))

    if not input_file or not os.path.exists(input_file):
        print("ERROR: No user metrics JSON file found. Run fetch_user_metrics.sh first.")
        sys.exit(1)

    print(f"Processing: {input_file}")

    # Load JSON data
    with open(input_file, 'r') as f:
        json_data = json.load(f)

    yesterday = datetime.now() - timedelta(days=1)

    # Extract user data, filter for past day activity, sort by username
    all_users = []
    for seat in json_data.get('seats', []):
        assignee = seat.get('assignee', {})
        last_activity = parse_datetime(seat.get('last_activity_at'))
        last_auth = parse_datetime(seat.get('last_authenticated_at'))
        created_at = parse_datetime(seat.get('created_at'))

        active_in_past_day = "Yes" if last_activity and last_activity > yesterday else "No"

        all_users.append({
            'Username': assignee.get('login', 'Unknown'),
            'PlanType': seat.get('plan_type', ''),
            'LastActivityAt': last_activity.isoformat() if last_activity else '',
            'LastActivityEditor': seat.get('last_activity_editor', ''),
            'LastAuthenticatedAt': last_auth.isoformat() if last_auth else '',
            'CreatedAt': created_at.isoformat() if created_at else '',
            'ActiveInPastDay': active_in_past_day
        })

    # Filter for past day and sort by username
    past_day_users = sorted(
        [u for u in all_users if u['ActiveInPastDay'] == 'Yes'],
        key=lambda x: x['Username'].lower()
    )

    print(f"Total users: {len(all_users)}")
    print(f"Users active in past day: {len(past_day_users)}")

    # Export to Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"user_metrics_pastday_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = 'ActiveUsers'

    if past_day_users:
        headers = list(past_day_users[0].keys())
        ws.append(headers)
        for user in past_day_users:
            ws.append([user[h] for h in headers])

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)

    print(f"Excel file created: {excel_path}")


if __name__ == '__main__':
    main()
