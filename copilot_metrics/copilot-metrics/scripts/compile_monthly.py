#!/usr/bin/env python3
"""
============================================
Compile Monthly Copilot Metrics
============================================
Combines all daily spreadsheets into one
monthly summary spreadsheet
Run at end of each month
============================================
"""

import argparse
import glob
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='Compile monthly Copilot metrics')
    parser.add_argument('--output-dir', default='./output', help='Output directory')
    parser.add_argument('--month', default=datetime.now().strftime("%Y-%m"), help='Month to compile (YYYY-MM)')
    args = parser.parse_args()

    output_dir = os.path.abspath(args.output_dir)
    month = args.month
    month_folder = os.path.join(output_dir, month)

    if not os.path.exists(month_folder):
        print(f"ERROR: Month folder not found: {month_folder}")
        sys.exit(1)

    print("============================================")
    print("Compiling Monthly Copilot Metrics")
    print("============================================")
    print(f"Month: {month}")
    print(f"Folder: {month_folder}")
    print("")

    # Find all daily spreadsheets for the month (exclude temp files and monthly summary)
    daily_files = sorted([
        f for f in glob.glob(os.path.join(month_folder, f"copilot_metrics_{month}-*.xlsx"))
        if not os.path.basename(f).startswith('~$') and '_monthly_summary' not in f
    ])

    if not daily_files:
        print(f"ERROR: No daily spreadsheets found in {month_folder}")
        sys.exit(1)

    print(f"Found {len(daily_files)} daily spreadsheet(s)")

    # Initialize collections for combined data
    all_summary = []
    all_active_users = []
    all_chat_requests = []
    all_lines_of_code = []

    for file_path in daily_files:
        print(f"  Processing: {os.path.basename(file_path)}")
        
        # Extract date from filename
        match = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(file_path))
        file_date = match.group(1) if match else "Unknown"

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)

            # Import Summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                rows = list(ws.values)
                if len(rows) > 1:
                    headers = rows[0]
                    for row in rows[1:]:
                        data = dict(zip(headers, row))
                        data['ReportDate'] = file_date
                        all_summary.append(data)

            # Import ActiveUsers sheet
            if 'ActiveUsers' in wb.sheetnames:
                ws = wb['ActiveUsers']
                rows = list(ws.values)
                if len(rows) > 1:
                    headers = rows[0]
                    for row in rows[1:]:
                        data = dict(zip(headers, row))
                        data['ReportDate'] = file_date
                        all_active_users.append(data)

            # Import ChatRequests sheet
            if 'ChatRequests' in wb.sheetnames:
                ws = wb['ChatRequests']
                rows = list(ws.values)
                if len(rows) > 1:
                    headers = rows[0]
                    for row in rows[1:]:
                        data = dict(zip(headers, row))
                        data['ReportDate'] = file_date
                        all_chat_requests.append(data)

            # Import LinesOfCode sheet
            if 'LinesOfCode' in wb.sheetnames:
                ws = wb['LinesOfCode']
                rows = list(ws.values)
                if len(rows) > 1:
                    headers = rows[0]
                    for row in rows[1:]:
                        data = dict(zip(headers, row))
                        data['ReportDate'] = file_date
                        all_lines_of_code.append(data)

            wb.close()
        except Exception as e:
            print(f"    WARNING: Error processing {os.path.basename(file_path)}: {e}")

    # Create monthly summary statistics
    print("")
    print("Creating monthly summary...")

    # Aggregate unique active users across the month
    unique_usernames = set()
    for user in all_active_users:
        if user.get('ActiveInPeriod') == 'Yes':
            unique_usernames.add(user.get('Username', ''))

    # Unique chat days (deduplicate by date)
    seen_chat_dates = set()
    unique_chat_days = []
    for chat in all_chat_requests:
        date = chat.get('Date', '')
        if date and date not in seen_chat_dates:
            seen_chat_dates.add(date)
            unique_chat_days.append(chat)

    # Unique lines data (deduplicate by date + language)
    seen_lines = set()
    unique_lines_data = []
    for line in all_lines_of_code:
        key = (line.get('Date', ''), line.get('Language', ''))
        if key not in seen_lines:
            seen_lines.add(key)
            unique_lines_data.append(line)

    # Calculate month totals
    month_total_chats = sum(int(d.get('TotalChatRequests', 0) or 0) for d in unique_chat_days)
    month_loc_added = sum(int(d.get('loc_added_sum', 0) or 0) for d in unique_lines_data)
    month_loc_suggested = sum(int(d.get('loc_suggested_to_add_sum', 0) or 0) for d in unique_lines_data)
    
    active_users_list = [int(d.get('TotalActiveUsers', 0) or 0) for d in unique_chat_days]
    month_avg_active_users = round(sum(active_users_list) / len(active_users_list), 1) if active_users_list else 0
    
    month_acceptance_rate = f"{round((month_loc_added / month_loc_suggested) * 100, 1)}%" if month_loc_suggested > 0 else "0%"

    monthly_summary = [
        {'Metric': 'Month', 'Value': month},
        {'Metric': 'Daily Reports Compiled', 'Value': len(daily_files)},
        {'Metric': 'Unique Active Users', 'Value': len(unique_usernames)},
        {'Metric': 'Avg Daily Active Users', 'Value': month_avg_active_users},
        {'Metric': 'Total Chat Requests', 'Value': month_total_chats},
        {'Metric': 'Total loc_added_sum', 'Value': month_loc_added},
        {'Metric': 'Total loc_suggested_to_add_sum', 'Value': month_loc_suggested},
        {'Metric': 'Monthly Acceptance Rate', 'Value': month_acceptance_rate},
    ]

    # User activity summary for the month
    user_activity = {}
    for user in all_active_users:
        username = user.get('Username', '')
        if not username:
            continue
        if username not in user_activity:
            user_activity[username] = {
                'Username': username,
                'DaysActive': 0,
                'LastActivityAt': '',
                'LastActivityEditor': ''
            }
        if user.get('ActiveInPeriod') == 'Yes':
            user_activity[username]['DaysActive'] += 1
        # Update last activity if more recent
        last_activity = user.get('LastActivityAt', '')
        if last_activity and last_activity > user_activity[username]['LastActivityAt']:
            user_activity[username]['LastActivityAt'] = last_activity
            user_activity[username]['LastActivityEditor'] = user.get('LastActivityEditor', '')

    user_monthly_summary = sorted(user_activity.values(), key=lambda x: x['Username'].lower())

    # Export to monthly summary file
    monthly_excel_path = os.path.join(month_folder, f"copilot_metrics_{month}_monthly_summary.xlsx")

    wb = Workbook()

    # Monthly Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'MonthlySummary'
    ws_summary.append(['Metric', 'Value'])
    for row in monthly_summary:
        ws_summary.append([row['Metric'], row['Value']])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # User Activity sheet
    ws_users = wb.create_sheet('UserActivity')
    if user_monthly_summary:
        headers = list(user_monthly_summary[0].keys())
        ws_users.append(headers)
        for user in user_monthly_summary:
            ws_users.append([user[h] for h in headers])
        for cell in ws_users[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    # Daily Chat Totals sheet
    ws_chat = wb.create_sheet('DailyChatTotals')
    if unique_chat_days:
        headers = ['Date', 'TotalChatRequests', 'IDEChatUsers', 'DotcomChatUsers', 'TotalActiveUsers', 'ChatsPerActiveUser']
        ws_chat.append(headers)
        for chat in sorted(unique_chat_days, key=lambda x: x.get('Date', '')):
            ws_chat.append([chat.get(h, '') for h in headers])
        for cell in ws_chat[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

    # Daily Code Totals sheet
    ws_lines = wb.create_sheet('DailyCodeTotals')
    if unique_lines_data:
        headers = ['Date', 'Editor', 'Language', 'loc_added_sum', 'loc_deleted_sum', 'loc_suggested_to_add_sum', 'agent_edit', 'Acceptances', 'Suggestions', 'AcceptanceRate']
        ws_lines.append(headers)
        for line in sorted(unique_lines_data, key=lambda x: (x.get('Date', ''), x.get('Language', ''))):
            ws_lines.append([line.get(h, '') for h in headers])
        for cell in ws_lines[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

    # Auto-width columns
    for ws in [ws_summary, ws_users, ws_chat, ws_lines]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(monthly_excel_path)

    print("")
    print("============================================")
    print("Monthly Compilation Complete")
    print("============================================")
    print(f"Output: {monthly_excel_path}")
    print("")
    print("Sheets included:")
    print(f"  1. MonthlySummary - Aggregated metrics for {month}")
    print("  2. UserActivity - Per-user activity summary")
    print("  3. DailyChatTotals - Chat requests by day")
    print("  4. DailyCodeTotals - Lines of code by day")


if __name__ == '__main__':
    main()
